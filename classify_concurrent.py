"""
Command-line tool to classify radiology reports using gemini-2.0-flash-001 API
and build a per-term confusion matrix.

Uses a .env file for the GEMINI_API_KEY.
Uses config.json for settings (terms, columns, model) and
prompt_config.json for the system prompt.

Usage:
    python prompt.py <excel_file_path>

This script performs the following steps:
1. Loads configuration from config.json and prompt_config.json.
2. Dynamically builds the system prompt, injecting the term list.
3. Loads the GEMINI_API_KEY from the .env file.
4. Reads the input Excel file provided as a command-line argument.
5. Iterates through each row, calling the Gemini API to classify findings
   CONCURRENTLY using a ThreadPoolExecutor.
6. Saves these classifications to an intermediate Excel file (Radiology_Classification_Output.xlsx).
7. Computes a per-term confusion matrix from the classified data.
8. Generates a final, multi-sheet Excel report (<input_stem>_confusion_matrix.xlsx)
   with "Confusion Matrix" and "Positive" (detailed) sheets.
"""

import pandas as pd
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Sequence
from pathlib import Path
import logging
import argparse
import os
import json
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
import concurrent.futures  # NEW: Import for concurrent execution

# ---------------------------------------------------------------------------
# Configuration Loader
# ---------------------------------------------------------------------------


def load_configuration(
    config_file: str = "config.json", prompt_file: str = "prompt_config.json"
) -> dict:
    """
    Loads main JSON config and prompt JSON config, then dynamically
    builds the system prompt by injecting the term list.

    Args:
        config_file (str): Filename for the main configuration JSON.
        prompt_file (str): Filename for the prompt configuration JSON.

    Returns:
        dict: A dictionary containing all loaded configuration and the
              dynamically built 'system_prompt'.

    Raises:
        FileNotFoundError: If config files are not found.
        json.JSONDecodeError: If config files are not valid JSON.
        KeyError: If prompt_config.json is missing required sections.
    """
    # Get the directory where the script is located
    script_dir = Path(__file__).parent.resolve()
    config_path = script_dir / config_file
    prompt_path = script_dir / prompt_file

    config = {}
    prompt_config = {}

    # Load main config (config.json)
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        logging.info(f"Loaded configuration from {config_file}")
    except FileNotFoundError:
        logging.error(f"FATAL: Main config file not found at {config_path}")
        raise
    except json.JSONDecodeError:
        logging.error(f"FATAL: Could not parse {config_file}. Check for JSON errors.")
        raise
    except Exception as e:
        logging.error(f"FATAL: Could not load {config_file}. Error: {e}")
        raise

    # Load prompt config (prompt_config.json)
    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt_config = json.load(f)
        logging.info(f"Loaded prompt configuration from {prompt_file}")
    except FileNotFoundError:
        logging.error(f"FATAL: Prompt config file not found at {prompt_path}")
        raise
    except json.JSONDecodeError:
        logging.error(f"FATAL: Could not parse {prompt_file}. Check for JSON errors.")
        raise
    except Exception as e:
        logging.error(f"FATAL: Could not load {prompt_file}. Error: {e}")
        raise

    # --- DYNAMICALLY BUILD THE SYSTEM PROMPT ---

    # 1. Get the list of terms from the main config
    terms_list = config.get("terms", [])
    if not terms_list:
        logging.warning(f"No 'terms' found in {config_file}")

    # 2. Join each prompt section (from arrays) using newlines
    try:
        # These keys must exist in prompt_config.json
        role_def = "\n".join(prompt_config["role_definition"])
        term_header = "\n".join(prompt_config["terminology_header"])
        instructions = "\n".join(prompt_config["instructions"])
        examples = "\n".join(prompt_config["examples"])
        output_rules = "\n".join(prompt_config["output_rules"])
    except KeyError as e:
        logging.error(f"FATAL: Missing expected key {e} in {prompt_file}.")
        logging.error(
            "Please ensure prompt_config.json contains: role_definition, terminology_header, instructions, examples, output_rules."
        )
        raise
    except TypeError as e:
        logging.error(
            f"FATAL: Error processing {prompt_file}. Are all prompt values (e.g., 'instructions') arrays of strings? Error: {e}"
        )
        raise

    # 3. Create the term list string (one term per line)
    terms_string = "\n".join(terms_list)

    # 4. Assemble the final prompt from all parts
    system_prompt_parts = [
        role_def,
        term_header,
        terms_string,  # <-- Injects the dynamic list here
        "",  # <-- Adds a blank line for spacing
        instructions,
        examples,
        output_rules,
    ]

    # 5. Add the final, assembled prompt string to the config dict
    # We join the major sections with a double newline for readability in the prompt
    config["system_prompt"] = "\n\n".join(system_prompt_parts)

    return config


# ---------------------------------------------------------------------------
# Load configuration & constants
# ---------------------------------------------------------------------------

# Setup basic logging config
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)

# Load all configuration from files at startup
try:
    CONFIG = load_configuration()
except Exception as e:
    # If config fails, we can't run. Exit the script.
    logging.critical(f"Failed to load configuration. Exiting. Error: {e}")
    sys.exit(1)  # Stop the script if config files are missing

# Load environment variables (e.g., GEMINI_API_KEY) from .env file
load_dotenv()

# --- Assign loaded config values to global constants ---

# List of medical terms to classify
TERMS: List[str] = CONFIG["terms"]

# Required output columns for the final confusion matrix sheet
REQUIRED_OUTPUT_COLS: List[str] = CONFIG["required_output_cols"]

# Standard column names for classification results
FN_COL = "fn_Positive"
FP_COL = "fp_Positive"
TN_COL = "tn_Positive"
TP_COL = "tp_Positive"

# Required input columns from the source Excel
REQUIRED_INPUT_COLS = CONFIG["required_input_cols"]

# Gemini model name
GEMINI_MODEL = CONFIG["gemini_model"]

# Intermediate output file name
INTERMEDIATE_OUTPUT = CONFIG["intermediate_output_file"]

# The complete, dynamically-built system prompt
SYSTEM_PROMPT = CONFIG["system_prompt"]

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class Counts:
    """
    A simple data structure to hold the aggregate counts (TP, FP, FN, TN)
    for a single medical term.
    """

    tp: int = 0
    fp: int = 0
    fn: int = 0
    tn: int = 0


# ---------------------------------------------------------------------------
# Parsing utilities
# ---------------------------------------------------------------------------


def parse_term_list(text: Optional[str]) -> set[str]:
    """
    Parse a string of terms separated by newlines, spaces, or commas into a
    clean set of strings.

    Args:
        text (Optional[str]): The input string.

    Returns:
        set[str]: A set of cleaned terms. Returns an empty set if input is None or empty.
    """
    if not isinstance(text, str) or not text.strip():
        return set()
    # Split by newline, space, or comma
    terms = re.split(r"\n|\s+|,", text.strip())
    # Return a set of stripped, non-empty terms
    return {t.strip() for t in terms if t.strip()}


# ---------------------------------------------------------------------------
# Gemini API integration for classification
# ---------------------------------------------------------------------------


def classify_with_gemini(
    idx: int, ai_findings: str, rad_findings: str
) -> Tuple[int, Dict[str, str], Optional[str]]:
    """
    Use the Gemini API to classify a single case into FN/FP/TN/TP.

    Args:
        idx (int): The row index of the classification (for tracking).
        ai_findings (str): The text from the "Findings (AI report)" column.
        rad_findings (str): The text from the "Findings (original radiologist report)" column.

    Returns:
        Tuple[int, Dict[str, str], Optional[str]]:
            (row_index, classifications_dict, error_message)
            classifications_dict is non-empty on success, error_message is non-empty on failure.

    Raises:
        ValueError: If API key is missing. (Handled internally now).
    """
    # Get API key from environment
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return idx, {}, "GEMINI_API_KEY not found in environment variables."

    # Configure the client
    genai.configure(api_key=api_key)

    # Configure the model, setting response type to JSON
    # This must be done on every thread if the client is not thread-safe or for configuration changes
    model = genai.GenerativeModel(
        GEMINI_MODEL,
        generation_config=genai.GenerationConfig(response_mime_type="application/json"),
    )

    # Create the user-specific prompt with the findings for this row
    user_prompt = f"""
AI Findings: “{ai_findings}”

Radiologist Findings: “{rad_findings}”
"""

    try:
        # Send the prompt (system + user) to the API
        response = model.generate_content([SYSTEM_PROMPT, user_prompt])

        # Clean up the response, removing markdown backticks if present
        response_text = response.text.strip()
        response_text = re.sub(r"^```json\s*|\s*```$", "", response_text)

        # Parse the cleaned text as JSON
        result = json.loads(response_text)

        # Validate the JSON structure to ensure all keys are present
        expected_keys = {"fn_Positive", "fp_Positive", "tn_Positive", "tp_Positive"}
        if set(result.keys()) != expected_keys:
            error = f"Invalid JSON structure from Gemini: {result}"
            return idx, {}, error

        # Success: Return index, results, and no error message
        return idx, result, None

    except json.JSONDecodeError:
        # Handle cases where the model output is not valid JSON
        error = f"Failed to parse Gemini response as JSON: {response.text[:100]}..."
        return idx, {}, error
    except Exception as e:
        # Handle other errors, especially content filtering
        if "response" in locals() and hasattr(response, "prompt_feedback"):
            error = f"Gemini API request blocked: {response.prompt_feedback}"
        else:
            # Re-raise any other exception
            error = f"Gemini API error: {str(e)}"
        return idx, {}, error


# ---------------------------------------------------------------------------
# Generate intermediate classification Excel (NOW CONCURRENT)
# ---------------------------------------------------------------------------


def generate_classification_excel(df: pd.DataFrame, file_path: Path) -> pd.DataFrame:
    """
    Generate an intermediate Excel file with classifications using Gemini API.
    Utilizes concurrent.futures.ThreadPoolExecutor to process API calls in parallel.

    Args:
        df (pd.DataFrame): The input DataFrame from the source Excel.
        file_path (Path): The path to the *original* input Excel file.

    Returns:
        pd.DataFrame: The DataFrame with new classification columns added.
    """
    logging.info("Generating intermediate classification Excel (concurrently)...")
    classified_df = df.copy()

    # Ensure classification columns exist, even if empty
    for col in [FN_COL, FP_COL, TN_COL, TP_COL]:
        if col not in classified_df.columns:
            classified_df[col] = ""

    # Define max workers for concurrency (a common practice is to use 10-20 for API calls)
    MAX_WORKERS = 10
    futures = []

    # Use ThreadPoolExecutor for concurrent synchronous API calls
    logging.info(f"Starting concurrent classification with max {MAX_WORKERS} workers...")

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 1. Submit all rows to the thread pool
        for idx, row in classified_df.iterrows():
            ai_findings = row.get("Findings (AI report)", "")
            rad_findings = row.get("Findings (original radiologist report)", "")

            # Handle potential NaN or non-string inputs from Excel
            ai_findings = "" if pd.isna(ai_findings) else str(ai_findings)
            rad_findings = "" if pd.isna(rad_findings) else str(rad_findings)

            # Submit the task to the pool
            future = executor.submit(
                classify_with_gemini, idx, ai_findings, rad_findings
            )
            futures.append(future)

        # 2. Process results as they complete
        for future in concurrent.futures.as_completed(futures):
            try:
                # The result is the tuple: (idx, classifications, error_message)
                idx, classifications, error_msg = future.result()

                # Safely get CaseID for better logging
                case_id = classified_df.at[idx, "CaseID"]

                if error_msg:
                    # Update row with failure status
                    logging.warning(
                        f"Failed to classify CaseID: {case_id} (Row {idx}). Error: {error_msg}"
                    )
                    # Mark all classification columns as failed
                    classified_df.at[idx, FN_COL] = "CLASSIFICATION_ERROR"
                    classified_df.at[idx, FP_COL] = "CLASSIFICATION_ERROR"
                    classified_df.at[idx, TN_COL] = "CLASSIFICATION_ERROR"
                    classified_df.at[idx, TP_COL] = "CLASSIFICATION_ERROR"
                else:
                    # Update row with successful classification
                    classified_df.at[idx, FN_COL] = (
                        "\n".join(classifications["fn_Positive"].split(","))
                        if classifications["fn_Positive"]
                        else ""
                    )
                    classified_df.at[idx, FP_COL] = (
                        "\n".join(classifications["fp_Positive"].split(","))
                        if classifications["fp_Positive"]
                        else ""
                    )
                    classified_df.at[idx, TN_COL] = (
                        "\n".join(classifications["tn_Positive"].split(","))
                        if classifications["tn_Positive"]
                        else ""
                    )
                    classified_df.at[idx, TP_COL] = (
                        "\n".join(classifications["tp_Positive"].split(","))
                        if classifications["tp_Positive"]
                        else ""
                    )
                    logging.info(
                        f"Classified CaseID: {case_id} (Row {idx}) successfully."
                    )

            except Exception as e:
                # Catches potential issues within the future object itself
                logging.critical(
                    f"A critical error occurred while getting a result from the thread pool: {e}"
                )

    # Define the path for the intermediate output file in the same directory
    intermediate_path = file_path.parent / INTERMEDIATE_OUTPUT

    # --- Format and save the intermediate file ---
    try:
        with pd.ExcelWriter(intermediate_path, engine="openpyxl") as writer:
            # Write the DataFrame to the 'Classifications' sheet
            classified_df.to_excel(writer, index=False, sheet_name="Classifications")

            # Get the worksheet object for formatting
            worksheet = writer.sheets["Classifications"]

            # Define styles for formatting
            header_font = Font(bold=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            wrap_alignment = Alignment(wrap_text=True, vertical="top")

            # Apply bold headers and borders
            for cell in worksheet[1]:
                cell.font = header_font
                cell.border = border

            # Apply wrapping and borders to all data cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = wrap_alignment

            # Auto-adjust column widths (up to a max of 60)
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        # Handle multi-line cells to get max width
                        if cell.value:
                            lines = str(cell.value).split("\n")
                            # Find the longest line in the cell
                            cell_len = max(len(line) for line in lines if line)
                            if not cell_len:
                                cell_len = 0
                        else:
                            cell_len = 0
                        # Clamp width to a max of 60
                        max_length = min(max(cell_len, max_length), 60)
                    except:
                        pass  # Ignore errors on special cell types
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width

            # Apply an Excel table style for better readability and filtering
            tab = Table(displayName="ClassificationTable", ref=worksheet.dimensions)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(tab)

        logging.info(f"Intermediate Excel saved: {intermediate_path}")
    except Exception as e:
        logging.error(f"Failed to write intermediate Excel file: {e}")
        # Continue anyway, as classified_df is in memory

    return classified_df


# ---------------------------------------------------------------------------
# Confusion matrix computation
# ---------------------------------------------------------------------------


def compute_confusion_matrix(
    df: pd.DataFrame,
    terms: Sequence[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compute per-term confusion matrix using the pre-classified columns
    from the intermediate DataFrame.

    Args:
        df (pd.DataFrame): The classified DataFrame (output of generate_classification_excel).
        terms (Sequence[str]): The list of medical terms to analyze.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]:
            1. confusion_df: DataFrame with aggregate TP/FP/FN/TN and metrics per term.
            2. per_row_df: DataFrame with per-row, per-term AI vs. GT classifications.
    """
    # Dictionary to store aggregate counts per term
    counts: Dict[str, Counts] = {t: Counts() for t in terms}
    # List to store per-row, per-term classifications (for 'Positive' sheet)
    detail_rows: List[Dict[str, object]] = []

    # Iterate over the *classified* DataFrame
    for idx, row in df.iterrows():
        # Parse the newline-separated terms back into sets
        fn_set = parse_term_list(row.get(FN_COL))
        fp_set = parse_term_list(row.get(FP_COL))
        tn_set = parse_term_list(row.get(TN_COL))
        tp_set = parse_term_list(row.get(TP_COL))

        detail = {"RowIndex": idx}
        # Check each term against the sets to categorize it (TP, FP, FN, TN)
        for term in terms:
            if term in tp_set:
                counts[term].tp += 1
                ai_pos = 1
                gt_pos = 1
            elif term in fp_set:
                counts[term].fp += 1
                ai_pos = 1
                gt_pos = 0
            elif term in fn_set:
                counts[term].fn += 1
                ai_pos = 0
                gt_pos = 1
            elif term in tn_set:
                counts[term].tn += 1
                ai_pos = 0
                gt_pos = 0
            else:
                # Default to True Negative if not mentioned anywhere by the API
                # This assumes missing terms are negative for both
                counts[term].tn += 1
                ai_pos = 0
                gt_pos = 0

            # Store the per-row AI vs. Ground Truth (GT) state
            detail[f"AI_{term}"] = ai_pos
            detail[f"GT_{term}"] = gt_pos
        detail_rows.append(detail)

    # Now, aggregate the counts into a new DataFrame
    records: List[Dict[str, object]] = []
    for term in terms:
        c = counts[term]
        support = c.tp + c.fn
        predicted_pos = c.tp + c.fp
        total = c.tp + c.fp + c.fn + c.tn

        # Calculate metrics, handling division by zero
        sensitivity = (c.tp / (c.tp + c.fn)) if (c.tp + c.fn) > 0 else 0.0
        specificity = (c.tn / (c.tn + c.fp)) if (c.tn + c.fp) > 0 else 0.0
        agreement = ((c.tp + c.tn) / total) if total > 0 else 0.0

        records.append(
            {
                "Term": term,
                "TP": c.tp,
                "FP": c.fp,
                "FN": c.fn,
                "TN": c.tn,
                "Support": support,
                "Predicted_Positive": predicted_pos,
                "Total": total,
                "Sensitivity": sensitivity,
                "Specificity": specificity,
                "Agreement": agreement,
            }
        )

    # Create the main confusion matrix DataFrame
    confusion_df = pd.DataFrame.from_records(records)
    # Create the DataFrame with per-row details
    per_row_df = pd.DataFrame.from_records(detail_rows)

    return confusion_df, per_row_df


def build_requested_output(confusion_df: pd.DataFrame) -> pd.DataFrame:
    """
    Transform the computed confusion matrix DataFrame into the final
    schema specified by REQUIRED_OUTPUT_COLS.

    Args:
        confusion_df (pd.DataFrame): The DataFrame from compute_confusion_matrix.

    Returns:
        pd.DataFrame: A DataFrame formatted to match the "Confusion Matrix" sheet spec.
    """
    df = confusion_df.copy()

    # Calculate statistics based on the matrix counts
    df["Positive Ground Truth"] = df["TP"] + df["FN"]
    df["Negative Ground Truth"] = df["TN"] + df["FP"]
    df["Check"] = df["TP"] + df["FP"] + df["FN"] + df["TN"]
    df["Ground Truth Check"] = df["Positive Ground Truth"] + df["Negative Ground Truth"]

    # Build the final DataFrame in the specified order
    out = pd.DataFrame(
        {
            "condition": df["Term"],
            "tp_Positive": df["TP"],
            "fn_Positive": df["FN"],
            "tn_Positive": df["TN"],
            "fp_Positive": df["FP"],
            "Sensitivity": df["Sensitivity"].round(4),
            "Specificity": df["Specificity"].round(4),
            "Check": df["Check"],
            "Positive Ground Truth": df["Positive Ground Truth"],
            "Negative Ground Truth": df["Negative Ground Truth"],
            "Ground Truth Check": df["Ground Truth Check"],
            "Radiologist Agreement Rate": df["Agreement"].round(4),
        }
    )
    # Insert the required blank column
    out.insert(11, "", "")

    # Ensure all required columns are present, even if empty
    for col in REQUIRED_OUTPUT_COLS:
        if col not in out.columns:
            out[col] = "" if col == "" else 0  # Add blank or 0

    # Reorder columns to match the exact specification
    out = out[REQUIRED_OUTPUT_COLS]
    return out


def build_details_df(
    input_df: pd.DataFrame, per_row_df: pd.DataFrame, terms: Sequence[str]
) -> pd.DataFrame:
    """
    Build the detailed "Positive" sheet by combining the original input data
    with the per-row classification results.

    Args:
        input_df (pd.DataFrame): The *original* input DataFrame.
        per_row_df (pd.DataFrame): The per-row classification data from compute_confusion_matrix.
        terms (Sequence[str]): The list of medical terms.

    Returns:
        pd.DataFrame: The DataFrame for the "Positive" sheet.
    """
    # Reset index to align with per_row_df's implicit 0-based index
    details_df = input_df.copy().reset_index(drop=True)
    per_row_df = per_row_df.copy().reset_index(drop=True)

    details_df["Original Radiologist"] = ""  # Placeholder column

    fn_lists, fp_lists, tp_lists, tn_lists = [], [], [], []

    # Safety check: ensure DataFrames have the same length
    if len(per_row_df) != len(details_df):
        logging.warning(
            f"Mismatch in row counts: input_df has {len(details_df)}, per_row_df has {len(per_row_df)}"
        )
        # Truncate to the shorter length to avoid errors
        min_len = min(len(per_row_df), len(details_df))
        per_row_df = per_row_df.iloc[:min_len]
        details_df = details_df.iloc[:min_len]

    # Iterate row by row to build string-based detail columns
    for i in range(len(per_row_df)):
        fn, fp, tp, tn = [], [], [], []

        # For each term, find its category (TP/FP/FN/TN)
        for term in terms:
            ai = per_row_df.at[i, f"AI_{term}"]
            gt = per_row_df.at[i, f"GT_{term}"]
            ai_val = "Positive" if ai else "Negative"
            gt_val = "Positive" if gt else "Negative"
            # Create a human-readable string: "pneumonia - (AI: Positive, Original: Negative)"
            val_str = f"{term} - (AI: {ai_val}, Original: {gt_val})"

            # Append the string to the correct category list
            if ai == 0 and gt == 1:
                fn.append(val_str)
            elif ai == 1 and gt == 0:
                fp.append(val_str)
            elif ai == 1 and gt == 1:
                tp.append(val_str)
            elif ai == 0 and gt == 0:
                tn.append(val_str)

        # Join all terms for that row into a single multi-line string
        fn_lists.append("\n".join(fn))
        fp_lists.append("\n".join(fp))
        tp_lists.append("\n".join(tp))
        tn_lists.append("\n".join(tn))

    # Add the new detail columns to the DataFrame
    details_df["False Negative"] = fn_lists
    details_df["False Positive"] = fp_lists
    details_df["True Positive"] = tp_lists
    details_df["True Negative"] = tn_lists

    # Add per-term AI classification ('Positive'/'Negative')
    for term in terms:
        details_df[f"{term}_AI"] = [
            "Positive" if per_row_df.at[i, f"AI_{term}"] else "Negative"
            for i in range(len(per_row_df))
        ]

    # Add placeholder column
    details_df["gpt_charge"] = 0.0

    # Add per-term category ('True Positive', 'False Positive', etc.)
    for term in terms:
        categories = []
        for i in range(len(per_row_df)):
            ai = per_row_df.at[i, f"AI_{term}"]
            gt = per_row_df.at[i, f"GT_{term}"]
            if ai == 1 and gt == 1:
                categories.append("True Positive")
            elif ai == 1 and gt == 0:
                categories.append("False Positive")
            elif ai == 0 and gt == 1:
                categories.append("False Negative")
            else:
                categories.append("True Negative")
        details_df[term] = categories

    # Add blank columns as requested
    details_df[""] = ""
    details_df[" "] = ""

    # Add per-term Ground Truth classification ('Positive'/'Negative')
    for term in terms:
        details_df[f"{term}_Original"] = [
            "Positive" if per_row_df.at[i, f"GT_{term}"] else "Negative"
            for i in range(len(per_row_df))
        ]

    return details_df


# ---------------------------------------------------------------------------
# I/O and CLI
# ---------------------------------------------------------------------------


def process_file(
    file_path: Path,
    terms: Sequence[str],
) -> Path:
    """
    Main processing pipeline for a single file.

    Args:
        file_path (Path): Path to the input Excel file.
        terms (Sequence[str]): List of medical terms to analyze.

    Returns:
        Path: The path to the final generated output Excel file.

    Raises:
        ValueError: If required columns are missing from the input file.
    """
    # 1. Read the input file
    logging.info("Reading input Excel: %s", file_path)
    df = pd.read_excel(file_path, engine="openpyxl")

    # 2. Validate input columns
    missing_cols = [col for col in REQUIRED_INPUT_COLS if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns in input Excel: {missing_cols}")

    # 3. Call Gemini API for each row and get classified DataFrame
    classified_df = generate_classification_excel(df, file_path)

    # 4. Compute the confusion matrix from classified data
    confusion_df, per_row_df = compute_confusion_matrix(classified_df, terms)

    # 5. Build the two output DataFrames for Excel
    requested_df = build_requested_output(confusion_df)
    # Note: 'classified_df' is used here, not the original 'df', to include
    # the intermediate FN/FP/TN/TP columns in the 'Positive' sheet.
    details_df = build_details_df(classified_df, per_row_df, terms)

    # 6. Define the final output path
    out_path = file_path.with_name(f"{file_path.stem}_confusion_matrix.xlsx")
    logging.info("Writing final output Excel: %s", out_path)

    # 7. Write both DataFrames to a two-sheet Excel file
    # All writing and formatting MUST happen inside the 'with' block
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Write data to respective sheets
        requested_df.to_excel(writer, index=False, sheet_name="Confusion Matrix")
        details_df.to_excel(writer, index=False, sheet_name="Positive")

        # --- Apply formatting to the 'Positive' sheet ---
        worksheet_details = writer.sheets["Positive"]
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # Apply text wrapping to all cells
        for row in worksheet_details.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment

        # Auto-adjust column widths for 'Positive' sheet
        for col in worksheet_details.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        cell_len = max(len(line) for line in lines if line)
                        if not cell_len:
                            cell_len = 0
                    else:
                        cell_len = 0
                    max_length = min(max(cell_len, max_length), 60)
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet_details.column_dimensions[column].width = adjusted_width

    # The 'with' block automatically saves the file upon exiting
    return out_path


def build_arg_parser() -> argparse.ArgumentParser:
    """
    Builds the command-line argument parser.

    Returns:
        argparse.ArgumentParser: The configured argument parser.
    """
    parser = argparse.ArgumentParser(
        description="Build confusion matrix from labeled Excel using Gemini API for classification."
    )
    parser.add_argument(
        "file_path",
        type=str,
        help="Path to input Excel (.xlsx) file",
    )
    return parser


def main() -> None:
    """
    Main entry point for the script.
    Parses arguments, checks file existence, and runs the processing pipeline.
    """
    # Set up and parse command-line arguments
    parser = build_arg_parser()
    args = parser.parse_args()

    # Resolve the input file path
    file_path = Path(args.file_path).expanduser().resolve()

    # Basic validation: check if file exists
    if not file_path.exists():
        logging.error(f"Input file not found: {file_path}")
        sys.exit(1)

    try:
        # Run the main processing pipeline
        out_path = process_file(
            file_path=file_path,
            terms=TERMS,
        )
        # Print success messages to console
        print(f"\nSuccessfully completed processing.")
        print(f"Intermediate file saved to: {file_path.parent / INTERMEDIATE_OUTPUT}")
        print(f"Final confusion matrix saved to: {out_path}")
    except Exception as exc:
        # Log any critical failures
        logging.error("Processing failed: %s", exc, exc_info=True)
        sys.exit(1)


# Standard Python script entry point
if __name__ == "__main__":
    main()
