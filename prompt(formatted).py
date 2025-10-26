"""
Command-line tool to classify radiology reports using gemini-2.0-flash-001 API and build a per-term confusion matrix.
Uses a .env file to load the GEMINI_API_KEY. Generates an intermediate Excel file with classifications before
computing the confusion matrix.

Usage:
    python prompt.py <excel_file_path>

Requirements:
- Input Excel must include "CaseID", "Link to AI report", "Link to Rad report",
  "Findings (original radiologist report)", "Findings (AI report)" columns.
- GEMINI_API_KEY must be set in a .env file in the project directory.
- Output includes:
  - Intermediate file: Radiology_Classification_Output.xlsx with classifications.
  - Final file: <input_stem>_confusion_matrix.xlsx with sheets "Confusion Matrix", "Positive".
- Works in VS Code on Windows or any platform with Python and dependencies.

Notes:
- Uses gemini-2.0-flash-001 for classification.
- Handles synonyms, negations, and context as per enhanced_prompt.txt.
- Generates intermediate Excel before computing confusion matrix.
"""

import pandas as pd
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Sequence
from pathlib import Path
import logging
import argparse
import os
import json
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration & constants
# ---------------------------------------------------------------------------

# Load environment variables from .env file
load_dotenv()

# List of medical terms to classify
TERMS: List[str] = [
    "perihilar_infiltrate",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pulmonary_nodules",
    "pleural_effusion",
    "rtm",
    "focal_caudodorsal_lung",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]

# Required output columns for the confusion matrix sheet
REQUIRED_OUTPUT_COLS: List[str] = [
    "condition",
    "tp_Positive",
    "fn_Positive",
    "tn_Positive",
    "fp_Positive",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
    "",
    "Radiologist Agreement Rate",
]

# Column names for classification results
FN_COL = "fn_Positive"
FP_COL = "fp_Positive"
TN_COL = "tn_Positive"
TP_COL = "tp_Positive"

# Required input columns
REQUIRED_INPUT_COLS = [
    "CaseID",
    "Link to AI report",
    "Link to Rad report",
    "Findings (original radiologist report)",
    "Findings (AI report)",
]

# Gemini model name (corrected to gemini-2.0-flash-001)
GEMINI_MODEL = "gemini-2.0-flash-001"

# Intermediate output file name
INTERMEDIATE_OUTPUT = "Radiology_Classification_Output.xlsx"

# System prompt (unchanged from provided script)
SYSTEM_PROMPT = """
You are an expert Radiologist specialized in medical AI report validation.
Your task is to analyze a single case from chest X-ray findings comparing AI-generated reports and radiologist reports.

You will receive the findings from the original radiologist report and the AI report.

Your goal is to perform an intelligent comparative analysis between the AI-generated and radiologist findings, and classify into True Positive, False Positive, True Negative, and False Negative based on the predefined medical terms.

---

### MEDICAL TERMINOLOGY LIST
Use this terminology list as reference. Also recognize synonymous or semantically similar phrases — e.g., “heart enlargement” ≈ “cardiomegaly”, “fluid in lungs” ≈ “pleural_effusion”, “bronchial thickening” ≈ “bronchitis”.

perihilar_infiltrate  
pneumonia  
bronchitis  
interstitial  
diseased_lungs  
hypo_plastic_trachea  
cardiomegaly  
pulmonary_nodules  
pleural_effusion  
rtm  
focal_caudodorsal_lung  
focal_perihilar  
pulmonary_hypoinflation  
right_sided_cardiomegaly  
pericardial_effusion  
bronchiectasis  
pulmonary_vessel_enlargement  
left_sided_cardiomegaly  
thoracic_lymphadenopathy  
esophagitis  
vhs_v2

---

### STEP-BY-STEP INSTRUCTIONS

1. Extract Text:
   - Read the provided “Findings (original radiologist report)” and “Findings (AI report)”.
   - Normalize text: lowercase, remove extra spaces, handle minor typos, and standardize common medical abbreviations (e.g., “CHF” → “cardiomegaly”, “ILD” → “interstitial lung disease”).

2. Understand Context:
   - Interpret the meaning — do not rely solely on word presence.
   - Detect negations: “no evidence of pneumonia” means pneumonia = negative.
   - Detect uncertainty: “mildly suggestive of” → treat as **positive but low confidence** (still count as positive).
   - Distinguish anatomical mentions that imply the same finding (e.g., “perihilar opacity” → “perihilar_infiltrate”).

3. Term Classification:
   For each terminology, determine:
   - AI_Positive → AI report suggests or confirms that finding.
   - Rad_Positive → Radiologist report confirms that finding.
   - If a finding is explicitly ruled out (“no signs of”, “absent”), mark as **Negative**.

4. Generate Confusion Categories:
   - True Positive (TP): AI and Radiologist both detect it.
   - True Negative (TN): Neither detects it.
   - False Positive (FP): AI detects but Radiologist does not.
   - False Negative (FN): Radiologist detects but AI does not.

5. Output:
   - fn_Positive (comma-separated terms)  
   - fp_Positive (comma-separated terms)  
   - tn_Positive (comma-separated terms)  
   - tp_Positive (comma-separated terms)

---

### EXAMPLES

**Example 1:**
AI: “
1.MILD BRONCHOINTERSTITIAL PULMONARY PATTERN; MILD DYNAMIC AIRWAY DISEASE; ATELECTASIS or less likely, pneumonia.
2.Based on these AI airway/pulmonary findings, the probability of this patient having clinically detectable respiratory signs is MEDIUM (50%)
3.Are respiratory signs present? If YES.
4.Are respiratory signs present? If NO.
5.The evaluation has come up positive for a pulmonary nodular pattern.
6.The cardiac silhouette size is normal.”  

Radiologist: “Three orthogonal thoracic radiographs dated 29th September 2024 are available for review. There are no previous radiographs available for comparison. 
Airway findings: A smoothly marginated soft tissue opacity is variably present overlying the dorsal aspect of the trachea at the thoracic inlet. This opacity reduces approximately 50% of the dorsoventral diameter of the trachea. The intrathoracic trachea is normal. Within the lung parenchyma there is a mixed bronchointerstitial opacification, predominantly caudal dorsally. The mainstem bronchi taper poorly. There is mild focal interstitial opacification in the left cranial lung lobe adjacent to the rib sections, however some obliquity is present. No associated osteolysis of the ribs is seen.
Cardiovascular findings: The cardiac silhouette is normal in shape, size and margination. The cranial and caudal pulmonary vasculature is normal. The caudal vena cava is normal. The aorta and mainstem pulmonary artery have a normal outline in the vd/dv l image.
Mediastinum and pleural space: There is some ventral pleural fat accumulation.
Musculoskeletal findings: No significant abnormalities are detected.
Included abdomen: No significant abnormalities are detected.”

→ Output JSON: {"fn_Positive": "rtm,focal_caudodorsal_lung,thoracic_lymphadenopathy", "fp_Positive": "pulmonary_nodules,pulmonary_hypoinflation", "tn_Positive": "perihilar_infiltrate,pneumonia,hypo_plastic_trachea,cardiomegaly,pleural_effusion,focal_perihilar,right_sided_cardiomegaly,pericardial_effusion,pulmonary_vessel_enlargement,left_sided_cardiomegaly,esophagitis,vhs_v2", "tp_Positive": "bronchitis,interstitial,diseased_lungs,bronchiectasis"}

**Example 2:**
AI: “
1.The cardiac silhouette size is normal.
2.There is no evidence of INTERSTITIAL or ALVEOLAR PATTERN, BRONCHITIS, PERIHILAR INFILTRATE, DYNAMIC AIRWAY COLLAPSE or BRONCHIECTASIS or PLEURAL EFFUSION.”  

Radiologist: “Study:
Thoracic radiography: three images dated September 29, 2024
A right lateral projection of the abdomen is also present in the study.
Findings:
The cardiac silhouette is normal in size and shape. The pulmonary vasculature is normal in size. There is a mild generalized bronchointerstitial pulmonary pattern. No pulmonary nodules or masses are present. On the VD view, there is a thin pleural fissure line between the right cranial and middle lung lobes. There is no intrathoracic lymphadenopathy. There is mild T 12-T 13 spondylosis deformans. There is a small soft tissue opaque nodule in the subcutaneous tissues of the caudodorsal cervical region
Comments:
Please the other submission for interpretation of the abdomen.”  

→ Output JSON: {"fn_Positive": "bronchitis,interstitial,diseased_lungs", "fp_Positive": "", "tn_Positive": "perihilar_infiltrate,pneumonia,hypo_plastic_trachea,cardiomegaly,pulmonary_nodules,pleural_effusion,rtm,focal_caudodorsal_lung,focal_perihilar,pulmonary_hypoinflation,right_sided_cardiomegaly,pericardial_effusion,pulmonary_vessel_enlargement,left_sided_cardiomegaly,thoracic_lymphadenopathy,esophagitis,vhs_v2", "tp_Positive": ""}

---

### OUTPUT RULES
- Use comma-separated terms in each field.
- If no terms in a category → empty string "".
- Ensure logical consistency (a term can only belong to one category per case).
- Account for context and synonyms (e.g., “mild cardiac enlargement” = “cardiomegaly”).
- Handle plural/singular forms and small spelling variations.
- Interpret partial phrases intelligently — e.g., “tracheal hypoplasia” = “hypo_plastic_trachea”.
- Always assume clinical language precision, not plain text pattern matching.

- Output ONLY the JSON object, nothing else: {"fn_Positive": "...", "fp_Positive": "...", "tn_Positive": "...", "tp_Positive": "..."}
"""

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class Counts:
    tp: int = 0
    fp: int = 0
    fn: int = 0
    tn: int = 0


# ---------------------------------------------------------------------------
# Parsing utilities
# ---------------------------------------------------------------------------


def parse_term_list(text: Optional[str]) -> set[str]:
    """Parse newline, space, or comma-separated terms into a set."""
    if not isinstance(text, str) or not text.strip():
        return set()
    terms = re.split(r"\n|\s+|,", text.strip())
    return {t.strip() for t in terms if t.strip()}


# ---------------------------------------------------------------------------
# Gemini API integration for classification
# ---------------------------------------------------------------------------


def classify_with_gemini(ai_findings: str, rad_findings: str) -> Dict[str, str]:
    """
    Use gemini-2.0-flash-001 to classify a single case into FN/FP/TN/TP.
    Returns a dict with keys: fn_Positive, fp_Positive, tn_Positive, tp_Positive.
    Each value is a comma-separated string of terms.
    """
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError(
            "GEMINI_API_KEY not found in .env file or environment variables."
        )
    genai.configure(api_key=api_key)

    model = genai.GenerativeModel(
        GEMINI_MODEL,
        generation_config=genai.GenerationConfig(response_mime_type="application/json"),
    )
    user_prompt = f"""
AI Findings: “{ai_findings}”

Radiologist Findings: “{rad_findings}”
"""

    try:
        response = model.generate_content([SYSTEM_PROMPT, user_prompt])
        # Strip any potential markdown code blocks as a fallback
        response_text = response.text.strip()
        response_text = re.sub(r"^```json\s*|\s*```$", "", response_text)
        result = json.loads(response_text)
        expected_keys = {"fn_Positive", "fp_Positive", "tn_Positive", "tp_Positive"}
        if set(result.keys()) != expected_keys:
            raise ValueError(f"Invalid JSON structure from Gemini: {result}")
        return result
    except json.JSONDecodeError:
        raise ValueError(f"Failed to parse Gemini response as JSON: {response.text}")
    except Exception as e:
        raise ValueError(f"Gemini API error: {str(e)}")


# ---------------------------------------------------------------------------
# Generate intermediate classification Excel
# ---------------------------------------------------------------------------


def generate_classification_excel(df: pd.DataFrame, file_path: Path) -> pd.DataFrame:
    """
    Generate an intermediate Excel file with classifications using Gemini API.
    Returns the classified DataFrame.
    """
    logging.info("Generating intermediate classification Excel...")
    classified_df = df.copy()

    # Ensure required columns exist
    for col in [FN_COL, FP_COL, TN_COL, TP_COL]:
        if col not in classified_df.columns:
            classified_df[col] = ""

    # Classify each row and convert commas to newlines
    for idx, row in classified_df.iterrows():
        ai_findings = row.get("Findings (AI report)", "")
        rad_findings = row.get("Findings (original radiologist report)", "")
        try:
            classifications = classify_with_gemini(ai_findings, rad_findings)
            # Convert comma-separated terms to newline-separated
            classified_df.at[idx, FN_COL] = (
                "\n".join(classifications["fn_Positive"].split(","))
                if classifications["fn_Positive"]
                else ""
            )
            classified_df.at[idx, FP_COL] = (
                "\n".join(classifications["fp_Positive"].split(","))
                if classifications["fp_Positive"]
                else ""
            )
            classified_df.at[idx, TN_COL] = (
                "\n".join(classifications["tn_Positive"].split(","))
                if classifications["tn_Positive"]
                else ""
            )
            classified_df.at[idx, TP_COL] = (
                "\n".join(classifications["tp_Positive"].split(","))
                if classifications["tp_Positive"]
                else ""
            )
            logging.info(f"Classified row {idx} successfully.")
        except Exception as e:
            logging.error(f"Error classifying row {idx}: {e}")
            classified_df.at[idx, FN_COL] = ""
            classified_df.at[idx, FP_COL] = ""
            classified_df.at[idx, TN_COL] = ""
            classified_df.at[idx, TP_COL] = ""

    # Save intermediate Excel with formatting
    intermediate_path = file_path.with_name(INTERMEDIATE_OUTPUT)
    writer = pd.ExcelWriter(intermediate_path, engine="openpyxl")
    classified_df.to_excel(writer, index=False, sheet_name="Classifications")

    # Access the openpyxl workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets["Classifications"]

    # Define styles
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Apply bold headers and borders
    for cell in worksheet[1]:
        cell.font = header_font
        cell.border = border

    # Apply text wrapping and borders to all cells
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_alignment

    # Auto-adjust column widths (up to max width of 60)
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                max_length = min(max(length, max_length), 60)
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

    # Apply table style
    tab = Table(displayName="ClassificationTable", ref=worksheet.dimensions)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(tab)

    # Save the workbook
    writer.close()
    logging.info(f"Intermediate Excel saved: {intermediate_path}")

    return classified_df


# ---------------------------------------------------------------------------
# Confusion matrix computation
# ---------------------------------------------------------------------------


def compute_confusion_matrix(
    df: pd.DataFrame,
    terms: Sequence[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compute per-term confusion matrix using pre-classified columns."""
    counts: Dict[str, Counts] = {t: Counts() for t in terms}
    detail_rows: List[Dict[str, object]] = []
    total_rows = len(df)

    for idx, row in df.iterrows():
        # Parse terms, handling newlines
        fn_set = parse_term_list(row.get(FN_COL))
        fp_set = parse_term_list(row.get(FP_COL))
        tn_set = parse_term_list(row.get(TN_COL))
        tp_set = parse_term_list(row.get(TP_COL))

        detail = {"RowIndex": idx}
        for term in terms:
            if term in tp_set:
                counts[term].tp += 1
                ai_pos = 1
                gt_pos = 1
            elif term in fp_set:
                counts[term].fp += 1
                ai_pos = 1
                gt_pos = 0
            elif term in fn_set:
                counts[term].fn += 1
                ai_pos = 0
                gt_pos = 1
            elif term in tn_set:
                counts[term].tn += 1
                ai_pos = 0
                gt_pos = 0
            else:
                counts[term].tn += 1
                ai_pos = 0
                gt_pos = 0
            detail[f"AI_{term}"] = ai_pos
            detail[f"GT_{term}"] = gt_pos
        detail_rows.append(detail)

    records: List[Dict[str, object]] = []
    for term in terms:
        c = counts[term]
        support = c.tp + c.fn
        predicted_pos = c.tp + c.fp
        total = c.tp + c.fp + c.fn + c.tn

        records.append(
            {
                "Term": term,
                "TP": c.tp,
                "FP": c.fp,
                "FN": c.fn,
                "TN": c.tn,
                "Support": support,
                "Predicted_Positive": predicted_pos,
            }
        )

    confusion_df = pd.DataFrame.from_records(records)
    per_row_df = pd.DataFrame.from_records(detail_rows)
    return confusion_df, per_row_df


def build_requested_output(confusion_df: pd.DataFrame) -> pd.DataFrame:
    """Transform to requested schema with specified formulas."""
    df = confusion_df.copy()
    df["Positive Ground Truth"] = df["TP"] + df["FN"]
    df["Negative Ground Truth"] = df["TN"] + df["FP"]
    df["Check"] = df["TP"] + df["FP"] + df["FN"] + df["TN"]
    df["Sensitivity"] = df.apply(
        lambda r: r["TP"] / (r["TP"] + r["FN"]) if (r["TP"] + r["FN"]) != 0 else 0.0,
        axis=1,
    )
    df["Specificity"] = df.apply(
        lambda r: r["TN"] / (r["TN"] + r["FP"]) if (r["TN"] + r["FP"]) != 0 else 0.0,
        axis=1,
    )
    df["Radiologist Agreement Rate"] = df.apply(
        lambda r: (r["TP"] + r["TN"]) / r["Check"] if r["Check"] != 0 else 0.0, axis=1
    )
    df["Ground Truth Check"] = df["Positive Ground Truth"] + df["Negative Ground Truth"]

    out = pd.DataFrame(
        {
            "condition": df["Term"],
            "tp_Positive": df["TP"],
            "fn_Positive": df["FN"],
            "tn_Positive": df["TN"],
            "fp_Positive": df["FP"],
            "Sensitivity": df["Sensitivity"].round(4),
            "Specificity": df["Specificity"].round(4),
            "Check": df["Check"],
            "Positive Ground Truth": df["Positive Ground Truth"],
            "Negative Ground Truth": df["Negative Ground Truth"],
            "Ground Truth Check": df["Ground Truth Check"],
            "Radiologist Agreement Rate": df["Radiologist Agreement Rate"].round(4),
        }
    )
    out.insert(11, "", "")
    out = out[REQUIRED_OUTPUT_COLS]
    return out


def build_details_df(
    input_df: pd.DataFrame, per_row_df: pd.DataFrame, terms: Sequence[str]
) -> pd.DataFrame:
    """Build details for 'Positive' sheet."""
    details_df = input_df.copy().reset_index(drop=True)
    details_df["Original Radiologist"] = ""

    fn_lists = []
    fp_lists = []
    tp_lists = []
    tn_lists = []
    for i in range(len(per_row_df)):
        fn = []
        fp = []
        tp = []
        tn = []
        for term in terms:
            ai = per_row_df[f"AI_{term}"][i]
            gt = per_row_df[f"GT_{term}"][i]
            ai_val = "Positive" if ai else "Negative"
            gt_val = "Positive" if gt else "Negative"
            val_str = f"{term} - (AI: {ai_val}, Original: {gt_val})"
            if ai == 0 and gt == 1:
                fn.append(val_str)
            elif ai == 1 and gt == 0:
                fp.append(val_str)
            elif ai == 1 and gt == 1:
                tp.append(val_str)
            elif ai == 0 and gt == 0:
                tn.append(val_str)
        fn_lists.append(" ".join(fn))
        fp_lists.append(" ".join(fp))
        tp_lists.append(" ".join(tp))
        tn_lists.append(" ".join(tn))

    details_df["False Negative"] = fn_lists
    details_df["False Positive"] = fp_lists
    details_df["True Positive"] = tp_lists
    details_df["True Negative"] = tn_lists

    for term in terms:
        details_df[f"{term}_AI"] = [
            "Positive" if per_row_df[f"AI_{term}"][i] else "Negative"
            for i in range(len(per_row_df))
        ]

    details_df["gpt_charge"] = 0.0

    for term in terms:
        categories = []
        for i in range(len(per_row_df)):
            ai = per_row_df[f"AI_{term}"][i]
            gt = per_row_df[f"GT_{term}"][i]
            if ai == 1 and gt == 1:
                categories.append("True Positive")
            elif ai == 1 and gt == 0:
                categories.append("False Positive")
            elif ai == 0 and gt == 1:
                categories.append("False Negative")
            else:
                categories.append("True Negative")
        details_df[term] = categories

    details_df[""] = ""
    details_df[" "] = ""

    for term in terms:
        details_df[f"{term}_Original"] = [
            "Positive" if per_row_df[f"GT_{term}"][i] else "Negative"
            for i in range(len(per_row_df))
        ]

    return details_df


# ---------------------------------------------------------------------------
# I/O and CLI
# ---------------------------------------------------------------------------


def process_file(
    file_path: Path,
    terms: Sequence[str],
) -> Path:
    """Process input, generate intermediate classification Excel, and write final output."""
    logging.info("Reading input Excel: %s", file_path)
    df = pd.read_excel(file_path, engine="openpyxl")

    # Validate input columns
    missing_cols = [col for col in REQUIRED_INPUT_COLS if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns in input Excel: {missing_cols}")

    # Generate intermediate classification Excel
    classified_df = generate_classification_excel(df, file_path)

    # Compute confusion matrix using classified DataFrame
    confusion_df, per_row_df = compute_confusion_matrix(classified_df, terms)
    requested_df = build_requested_output(confusion_df)
    details_df = build_details_df(classified_df, per_row_df, terms)

    # Write final output
    out_path = file_path.with_name(f"{file_path.stem}_confusion_matrix.xlsx")
    logging.info("Writing final output Excel: %s", out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        requested_df.to_excel(writer, index=False, sheet_name="Confusion Matrix")
        details_df.to_excel(writer, index=False, sheet_name="Positive")

    return out_path


def build_arg_parser() -> argparse.ArgumentParser:  # Build argument parser
    parser = argparse.ArgumentParser(
        description="Build confusion matrix from labeled Excel using Gemini API for classification."
    )
    parser.add_argument(
        "file_path",
        type=str,
        help="Path to input Excel (.xlsx) file",
    )
    return parser


def main() -> None:  # Main function
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
    )
    parser = build_arg_parser()
    args = parser.parse_args()

    file_path = Path(args.file_path).expanduser().resolve()

    try:
        out_path = process_file(
            file_path=file_path,
            terms=TERMS,
        )
        print(str(out_path))
    except Exception as exc:
        logging.error("Processing failed: %s", exc)
        raise


if __name__ == "__main__":
    main()
